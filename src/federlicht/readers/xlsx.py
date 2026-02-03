from __future__ import annotations

from pathlib import Path

try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None


def _xlsx_available() -> bool:
    return openpyxl is not None


def read_xlsx_text(
    xlsx_path: Path,
    max_chars: int,
    max_sheets: int = 0,
    start_sheet: int = 0,
    max_rows: int = 200,
    max_cols: int = 50,
) -> str:
    if not _xlsx_available():
        return "openpyxl is not installed. Cannot read XLSX."
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)  # type: ignore[attr-defined]
    try:
        sheet_names = wb.sheetnames
        start_sheet = max(0, start_sheet)
        end_sheet = len(sheet_names) if max_sheets <= 0 else min(len(sheet_names), start_sheet + max_sheets)
        parts: list[str] = []
        char_count = 0
        truncated_chars = False
        for idx in range(start_sheet, end_sheet):
            ws = wb[sheet_names[idx]]
            parts.append(f"[sheet] {ws.title}")
            char_count += len(parts[-1]) + 1
            if max_chars and max_chars > 0 and char_count >= max_chars:
                truncated_chars = True
                break
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count >= max_rows:
                    break
                values = ["" if value is None else str(value) for value in row[:max_cols]]
                if any(values):
                    line = "\t".join(values)
                    parts.append(line)
                    char_count += len(line) + 1
                    if max_chars and max_chars > 0 and char_count >= max_chars:
                        truncated_chars = True
                        break
                row_count += 1
            if truncated_chars:
                break
        text = "\n".join(parts)
        if truncated_chars or (max_chars and max_chars > 0 and len(text) > max_chars):
            text = text[:max_chars].rstrip() + "\n\n[note] XLSX truncated: increase --max-chars or use start_page."
        if end_sheet < len(sheet_names):
            text = (
                text.rstrip()
                + f"\n\n[note] XLSX truncated: sheets {start_sheet + 1}-{end_sheet} of {len(sheet_names)}. "
                "Increase --max-pages or use start_page to read more sheets."
            )
        return text
    finally:
        wb.close()
